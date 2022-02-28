from asyncore import write
import json
import os
from flask import Flask, jsonify, request
from flask_cors import CORS
import pandas as pd
import openpyxl

app = Flask(__name__)
CORS(app)

@app.route('/saveReadings', methods=['GET'])
def addData():
    print(50)
    request_data = request.get_json()
    data = [] 
    data.append(request_data)
    print(data)
    df_data = pd.DataFrame(data)
    path = os.getcwd()
    filename = 'data.xlsx'
    sheet_name = "lab"
    if (os.path.exists(filename)) :
        wb = openpyxl.load_workbook(filename)
        if(not (sheet_name in wb.sheetnames)):
            wb.create_sheet(sheet_name)
        ws = wb.active
        for cell_data in data:
            row=ws.max_row+1
            ws.cell(column=1, row=row, value=cell_data['SSID'])
            ws.cell(column=2, row=row, value=cell_data['Strength'])
            wb.save(filename)
    else:
        df_data.to_excel(filename, index=False, sheet_name=sheet_name)

    return jsonify(data)


if __name__ == "main":
    app.run(host="localhost", port=8433, debug=True)